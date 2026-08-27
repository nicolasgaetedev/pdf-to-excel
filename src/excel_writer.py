from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def crear_excel(resultados, ruta_salida):
    try:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Facturas"

        encabezados = [
            "Archivo",
            "N° Factura",
            "Empresa",
            "RUT",
            "Fecha",
            "Subtotal",
            "IVA",
            "Total",
        ]

        hoja.append(encabezados)

        # Encabezados en negrita
        for celda in hoja[1]:
            celda.font = Font(bold=True)

        # Agregar datos
        for resultado in resultados:
            hoja.append([
                resultado["archivo"],
                resultado["numero"],
                resultado["empresa"],
                resultado["rut"],
                resultado["fecha"],
                resultado["subtotal"],
                resultado["iva"],
                resultado["total"],
            ])

        # Congelar encabezados
        hoja.freeze_panes = "A2"

        # Activar filtros
        hoja.auto_filter.ref = hoja.dimensions

        # Formato monetario
        for fila in range(2, hoja.max_row + 1):
            hoja[f"F{fila}"].number_format = '$#,##0'
            hoja[f"G{fila}"].number_format = '$#,##0'
            hoja[f"H{fila}"].number_format = '$#,##0'

        # Ajustar automáticamente el ancho de columnas
        for columna in hoja.columns:
            largo_maximo = 0
            letra_columna = get_column_letter(columna[0].column)

            for celda in columna:
                if celda.value is not None:
                    largo_maximo = max(
                        largo_maximo,
                        len(str(celda.value))
                    )

            hoja.column_dimensions[letra_columna].width = largo_maximo + 2

        # Anchos mínimos para las columnas monetarias
        hoja.column_dimensions["F"].width = max(
            hoja.column_dimensions["F"].width,
            14,
        )

        hoja.column_dimensions["G"].width = max(
            hoja.column_dimensions["G"].width,
            14,
        )

        hoja.column_dimensions["H"].width = max(
            hoja.column_dimensions["H"].width,
            14,
        )

        libro.save(ruta_salida)

        return True

    except PermissionError:
        print(
            f"✗ No se pudo guardar {ruta_salida}. "
            "Comprueba que el archivo no esté abierto en Excel."
        )
        return False

    except Exception as error:
        print(f"✗ Error al generar el Excel: {error}")
        return False